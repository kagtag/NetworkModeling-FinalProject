# -*- coding:utf8 -*-
import matplotlib.pyplot as plt
import math
import openpyxl
import logging
import wangyiyun.Neteasebox.api as api


import numpy as np
import networkx as nx

import time
import random
import datetime
logging.basicConfig(level=logging.INFO)

class music_entity:

    def __init__(self,user_id,graph_size,d):
        """

        :param user_id:
        :param graph_size:number of nodes in this graph
        """
        self.user_id=user_id #the target user
        self.netease=api.NetEase()
        self.graph_size=graph_size #number of nodes in the graph
        self.d=d #prob of taking one of the adjacent edge

        #依照从上到下的顺序,递归调用
        self.user_info_collect={}#{user id：[playlist ids,歌单id...]}
        self.play_info_collect={}#{playlist id:{information 各种信息,包括内含歌曲id列表}}
        self.song_info_collect={}#{song id:{information各种信息}}
        self.artist_info_collect={}#{artist id:{information-去豆瓣}}

        self.net_influence={}

        self.G=nx.DiGraph()
        self.pagerank=''
        self.q=''

        self.result_dict={} #store evaluation result

    def turn_undirected(self):
        self.G=self.G.to_undirected()

    #调用以规定要对谁进行推荐，类似于一个state machine
    def rec_for(self,q,l,result_size):
        self.q=q
        self.l=l #number of steps
        self.result_size=result_size

    #########################################################
    #########################################################
    #########################################################3
    #read graph from the database

    #input graph information from database
    def read_graph(self,path):
        pass


    #########################################################
    #########################################################
    #########################################################
    # methods to construct the graph


    def user_graph_dfs(self,user_id):
        """using depth first search to construct the network, starting from
        the node refered by user_id

        :param user_id: starting node
        """

        user_id=int(user_id)
        #首先处理该用户下所有歌单的信息
        self.user_plays(user_id)

        #处理完成后构建网络
        for play_id in self.user_info_collect[user_id]:
            creator=int(self.play_info_collect[play_id]['creator'])
            if self.G.has_edge(user_id,creator) or user_id==creator:
                continue
            else:
                if self.G.number_of_nodes() > self.graph_size:
                    print(self.G.number_of_nodes())
                    return

                self.G.add_edge(user_id,creator)
                self.user_graph_dfs(creator)

    def user_graph_bfs(self,user_id):
        """using breadth first search to construct the network, starting from
        the node refered by user_id

        :param user_id:
        :return:
        """

        user_id = int(user_id)  # 换为整数
        tmp_que=[user_id]

        while self.G.number_of_nodes()<self.graph_size:
            self.user_plays(user_id)

            # 处理完成后构建网络
            for play_id in self.user_info_collect[user_id]:

                creator = int(self.play_info_collect[play_id]['creator']) #换为整数
                if self.G.has_edge(user_id, creator) or user_id == creator:
                    continue
                else:
                    if self.G.number_of_nodes() > self.graph_size:
                        print(self.G.number_of_nodes())
                        return
                    self.G.add_edge(user_id, creator,
                                    weight=math.log(float(self.play_info_collect[play_id]['subcount'])))
                    tmp_que.append(creator)
            tmp_que.pop(0)
            user_id=tmp_que[0]


    def save_graph(self):
        #save all the data of the graph
        workbook = openpyxl.Workbook()
        worksheet = workbook.get_active_sheet()

        for idx,item in enumerate(self.user_info_collect.items()):
            worksheet.cell(row=idx+1,column=1).value=item[0]
            worksheet.cell(row=idx + 1, column=2).value = str(item[1])

        worksheet=workbook.create_sheet()

        for idx,item in enumerate(self.play_info_collect.items()):
            worksheet.cell(row=idx+1,column=1).value=item[0]
            for iidx,fea in enumerate(['name','creator','createtime','playcount','subcount','songs']):
                worksheet.cell(row=idx+1,column=iidx+2).value=str(item[1].get(fea,''))

        # worksheet=workbook.create_sheet()

        # for idx,item in enumerate(self.)

        workbook.save("wangyiyun/music_data.xlsx")

        nx.draw(self.G,node_size=45)
        plt.savefig("wangyiyun/graph.png")
        plt.show()


    #########################################################
    #########################################################
    #########################################################
    # methods to extract information through the Netease api


    #给一个歌单id，并计算歌单的各种属性
    def play_stat(self,play_id):
        """compute the average popularity of the playlist

        :param play_id:
        :return:
        """
        if self.play_info_collect[play_id].get("popularity",'')!='':
            return

        tmp=[self.song_info_collect[song]["cmttotal"] for song in self.play_info_collect[play_id]]
        tmp=[1/math.log(x) for x in tmp]
        self.play_info_collect[play_id]['popularity']=sum(tmp)/len(tmp)


    def song_attribute(self,song_id):
        """collect information of a song

        :param song_id: id of the target song
        """
        if self.song_info_collect.get(song_id,'')!='':
            return #information of this song has already been collected
        self.song_info_collect[song_id] = {}
        song = self.netease.song_detail(song_id)[0]

        logging.info("information of song "+str(song_id))
        # name
        self.song_info_collect[song_id]['name'] = song['name']

        # artist
        tmp_artist_id=song['artists'][0]['id']
        self.song_info_collect[song_id]['artistid'] = tmp_artist_id

        #if the artist have not been seen before, store the information in the dictionary
        if self.artist_info_collect.get(tmp_artist_id,'')=='':
            self.artist_info_collect[tmp_artist_id] = {}
            self.artist_info_collect[tmp_artist_id]['name'] = song['artists'][0]['name']

        #number of comments
        cmt = self.netease.song_comments(song_id)
        self.song_info_collect[song_id]['cmttotal'] = cmt['total']



    #collect songs of a given playlist
    def collect_songs(self,play_id):
        """collect songs of a playlist

        :param play_id: the target playlist
        """

        #check if this playlist have been seen before
        if self.play_info_collect.get(play_id,'')=='':
            self.play_info_collect[play_id]={}

        #若已收集过歌曲，则不再执行
        if self.play_info_collect[play_id].get('songs','')!='':
            return

        logging.info("collect songs of playlist "+str(play_id))
        song_pool = []
        attempt=0
        while True:
            try:
                # get songs'id in this play list
                for song in self.netease.playlist_detail(play_id):
                    song_pool.append(song['id'])

                    # disable
                    # self.song_attribute(song['id'])
                break
            except:
                attempt+=1
                if attempt>=5:
                    logging.info(str(play_id)+" playlist failed")
                    break
                logging.info(str(play_id) + " playlist is empty")
                time.sleep(5)
                continue

        # try:
        #     # get songs'id in this play list
        #     for song in self.netease.playlist_detail(play_id):
        #         song_pool.append(song['id'])
        #
        #         #disable
        #         # self.song_attribute(song['id'])
        #
        # except:
        #     logging.info(str(play_id)+" playlist is empty")

        self.play_info_collect[play_id]['songs'] = song_pool

    #输入用户id，收集用户的歌单信息
    def user_plays(self,user_id):
        """collect information of playlists of one user

        :param user_id: id of the user
        """
        time.sleep(random.random()+1.5)
        user_id = int(user_id)

        while True:
            user_lists = self.netease.user_playlist(user_id,limit=60)#get 60 playlists at most
            try:
                logging.info(str(user_id)+" number of playlist "+str(len(user_lists)))
                break
            except:
                time.sleep(10)
                logging.info("cannot get information of user "+str(user_id))
                continue

        self.user_info_collect[user_id]=[]
        for item in user_lists:


            # get information about the play list

            # "songs I like" playlist
            if "喜欢的音乐" in item['name']:
                play_id = item['id']
                self.user_info_collect[user_id].append(play_id) #add to the playlists record of the user

                self.play_info_collect[play_id]={}
                self.play_info_collect[play_id]['name'] = item['name']
                # self.play_info_collect[play_id]['creator'] = int(item['creator']['userId'])

                self.play_info_collect[play_id]['createtime'] = datetime.datetime.fromtimestamp(
                    item['createTime'] / 1000.0)
                self.play_info_collect[play_id]['creator'] = int(item['creator']['userId'])  # 换为整数
                self.play_info_collect[play_id]['playcount'] = item['playCount']
                self.play_info_collect[play_id]['subcount'] = int(item['subscribedCount'])

                self.collect_songs(play_id)
                continue

            #ignore playlists that have less than 500 subscribers or created by user him/erself
            if item['subscribedCount']<500 or item['creator']==user_id:
                continue

            play_id = item['id']
            self.user_info_collect[user_id].append(play_id)

            if play_id not in self.play_info_collect.keys():
                self.play_info_collect[play_id]={}
                self.play_info_collect[play_id]['name'] = item['name']
                self.play_info_collect[play_id]['createtime']=datetime.datetime.fromtimestamp(item['createTime'] / 1000.0)
                self.play_info_collect[play_id]['creator'] = int(item['creator']['userId']) #换为整数
                self.play_info_collect[play_id]['playcount'] = item['playCount']
                self.play_info_collect[play_id]['subcount'] = int(item['subscribedCount']) #换为整数

                #collect songs of this playlist
                self.collect_songs(play_id)



    #########################################################
    #########################################################
    #########################################################3
    # Personalized Pagerank


    # 寻找一个用户的代表性歌单
    def like_play(self, user_id, mode=1):
        user_id = int(user_id)
        i_like = ''
        # 抽取我喜欢的音乐
        flag = False

        # 成功在已有歌单里找到
        for item in self.user_info_collect[user_id]:
            if "喜欢的音乐" in self.play_info_collect[item]['name']:
                flag = True
                logging.info(str(self.play_info_collect[item]['name']))
                self.collect_songs(item)
                i_like = set(self.play_info_collect[item]['songs'])

                break

        # 没有成功在已有歌单里找到
        if flag == False:
            logging.info(str(user_id) + "\'songs I like\' playlist does not exist")
            tmp = self.user_info_collect[user_id][0]
            self.collect_songs(tmp)
            i_like = set(self.play_info_collect[tmp]['songs'])

        if mode == 1:
            return i_like
        elif mode == 2:
            a_like = set()
            for item in i_like:
                self.song_attribute(item)
                tmp = self.song_info_collect[item].get('artistid', '')
                if tmp == '':
                    self.song_attribute(item)
                    tmp = self.song_info_collect[item].get('artistid', '')
                a_like.add(tmp)

            return list(a_like)

    # 基于喜爱歌单关系计算相似用户
    def relevant_users(self, user_id):
        user_id = int(user_id)
        similar_users = []
        nodes = self.G.nodes()

        i_like = set(self.like_play(user_id))

        tmp_neighbor = set(self.G.predecessors(user_id)) | set(self.G.successors(user_id))

        if len(i_like) < 10:  # too few songs in the "songs I like" playlist, return all his/er neighbors
            return tmp_neighbor

        # 基于喜爱歌单计算用户之间相似性
        for cand_id in nodes:
            logging.info("compute similarity between "+str(user_id)+" and "+str(cand_id))
            other_like = set(self.like_play(cand_id))
            if len(other_like) < 10:
                continue

            similarity = len(other_like & i_like) / (len(i_like) if len(i_like) <= len(other_like) else len(other_like))
            if similarity > 0.3:
                similar_users.append(cand_id)

        return similar_users if len(similar_users) > 1 else tmp_neighbor

    # query 为目标用户的id值,基于该id值计算相似用户集合
    def per_pagerank(self):
        # d=0.1
        nodes = self.G.nodes()
        nonodes = len(nodes)
        trans_mat = np.zeros((nonodes, nonodes))

        u_Q = self.relevant_users(self.q)

        # compute transition matrix
        for i in range(nonodes):
            u = nodes[i]

            for j in range(nonodes):

                v = nodes[j]

                if self.G.has_edge(u, v):
                    # *math.log(self.G[u][v]['weight'])
                    trans_mat[i][j] = (1 - self.d) * (1 / len(u_Q) if v in u_Q else 0) + self.d / self.G.in_degree(v)
                else:
                    trans_mat[i][j] = (1 - self.d) * (1 / len(u_Q) if v in u_Q else 0)

        p_old = [0.1] * nonodes  # 0.1 is the initial score of each node

        for i in range(20):
            p_old = np.dot(trans_mat, np.transpose(p_old))

        logging.info("succuss transition matrix")
        self.pagerank = [p_old, nodes]
        # return p_old,nodes

    #########################################################
    #########################################################
    #########################################################
    # BestCoverage

    # compute l-step expansion set
    def step_expansion_set(self, ids):
        if isinstance(ids, int):
            S = set([ids])
        else:
            S = set(ids)
        V = set(self.G.nodes())

        for v in V - S:

            for u in S:
                try:
                    tmp_l = nx.shortest_path_length(self.G, u, v)
                    logging.info("length of the shortest path is " + str(tmp_l))
                except:
                    tmp_l = 10000
                if tmp_l <= self.l:
                    S.add(v)
                    break
        return S

    def step_expansion_ratio(self, ids):
        ori = len(ids)
        exp = self.step_expansion_set(ids)
        return len(exp) / ori

    def norm_rel(self, ids):
        if self.pagerank == '':
            self.per_pagerank()
        S = set(ids)
        print('norm_rel', len(S), self.result_size)

        r_p_new, r_nodes = self.page_rec()

        tmp = 0
        for i in range(self.result_size):
            tmp += r_p_new[i]
        tmp1 = tmp / self.result_size

        tmp = 0
        for item in S:
            try:
                tmp += r_p_new[np.where(r_nodes == item)[0][0]]
            except:
                logging.info(str(item))
        tmp2 = tmp / len(S)

        return tmp2 / tmp1

    def exp_rel(self, ids):
        if self.pagerank == '':
            self.per_pagerank()  ##?????
        p_new, nodes = self.pagerank

        V = self.step_expansion_set(ids)

        tmp = 0
        for item in V:
            tmp += p_new[nodes.index(item)]
        return tmp

        ####
        # order=np.argsort(-p_new)
        # ranked_p=nodes[order]
        ####

    # 给定集合S和节点v
    def marginal_utility(self, v, S):
        if self.pagerank == '':
            self.per_pagerank()
        if len(S) == 0:
            return self.exp_rel(v)

        p_new, nodes = self.pagerank
        Vv = self.step_expansion_set(v) - self.step_expansion_set(S)
        tmp = 0
        for item in Vv:
            tmp += p_new[nodes.index(item)]
        return tmp

    def best_coverage(self):
        if self.pagerank == '':
            self.per_pagerank()
        result = []
        p_new, nodes = self.pagerank
        while len(result) < self.result_size:

            tmp_max = 0
            max_item = ''
            for item in nodes:
                tmp = self.marginal_utility(item, result)
                if tmp_max < tmp:
                    tmp_max = tmp
                    max_item = item

            result.append(max_item)
        return result


    ########################################
    #########################################
    ########################################
    #methods for recommendation and evaluation

    def random_rec(self):
        """random recommendation

        :return:
        """
        result=[]
        for i in range(self.result_size):
            nodes=self.G.nodes()
            result.append(nodes[random.randint(0,len(nodes)-1)])
        return result

    def page_rec(self):
        """personalized pagerank recommendation

        :return:
        """
        if self.pagerank=='':
            self.per_pagerank()
        p_new,nodes=self.pagerank

        p_new=np.array(p_new)
        nodes=np.array(nodes)
        rank=np.argsort(-p_new)
        p_new=p_new[rank]
        nodes=nodes[rank]
        return p_new,nodes

    #衡量推荐结果,在给定k和推荐对象的条件下
    def rec_stat(self,tar=''):
        """evaluation of recommendation result

        :param tar: target user
        :return:
        """
        self.result_dict['pagerank']={}
        self.result_dict['random']={}
        self.result_dict['best']={}

        if tar=='':
            tar=self.user_id
        p_new,nodes=self.page_rec()
        p_rec=nodes[0:self.result_size]

        r_rec=self.random_rec()
        b_rec=self.best_coverage()

        #relevance
        self.result_dict['pagerank']['rel']=self.norm_rel(p_rec)
        self.result_dict['random']['rel'] =self.norm_rel(r_rec)
        self.result_dict['best']['rel'] =self.norm_rel(b_rec)

        #diversity
        self.result_dict['pagerank']['exp'] = self.step_expansion_ratio(p_rec)
        self.result_dict['random']['exp'] = self.step_expansion_ratio(r_rec)
        self.result_dict['best']['exp'] = self.step_expansion_ratio(b_rec)

        #exp rel
        self.result_dict['pagerank']['er'] = self.exp_rel(p_rec)
        self.result_dict['random']['er'] = self.exp_rel(r_rec)
        self.result_dict['best']['er'] = self.exp_rel(b_rec)

        # print(result_dict)
        workbook=openpyxl.Workbook()
        worksheet=workbook.get_active_sheet()
        lf=['rel','exp','er']
        lm=['pagerank','random','best']
        for i in range(len(lm)):
            for j in range(len(lf)):
                worksheet.cell(row=i+1,column=j+1).value=self.result_dict[lm[i]][lf[j]]

        workbook.save('wangyiyun/stat'+str(self.result_size)+'.xlsx')





#static methods
def to_unweighted(G):
    new_G=nx.Graph()
    for edges in G.edges():
        new_G.add_edge(edges[0],edges[1])
    return new_G


def degree_hist(G):
    G=to_unweighted(G)

    degree_sequence = sorted(nx.degree(G).values(), reverse=True)  # degree sequence
    dmax = max(degree_sequence)

    plt.loglog(degree_sequence, 'b-', marker='o')
    plt.title("Degree rank plot")
    plt.ylabel("degree")
    plt.xlabel("rank")

    # draw graph in inset
    plt.axes([0.45, 0.45, 0.45, 0.45])
    Gcc = sorted(nx.connected_component_subgraphs(G), key=len, reverse=True)[0]
    pos = nx.spring_layout(Gcc)
    plt.axis('off')
    nx.draw_networkx_nodes(Gcc, pos, node_size=20)
    nx.draw_networkx_edges(Gcc, pos, alpha=0.4)

    plt.savefig("wangyiyun/degree_histogram.png")
    plt.show()


def graph_stat(G):
    # print(G)
    G=to_unweighted(G)
    print("图为")
    print(G.edges())
    workbook=openpyxl.Workbook()
    worksheet=workbook.get_active_sheet()

    worksheet.cell(row=1,column=1).value=nx.average_clustering(G)
    worksheet.cell(row=1,column= 2).value =nx.degree_assortativity_coefficient(G)
    worksheet.cell(row=1, column=3).value =nx.diameter(G)
    worksheet.cell(row=1, column=4).value =nx.average_shortest_path_length(G)

    workbook.save("wangyiyun/new/stat.xlsx")


test_id="48548007"
graph=music_entity(user_id=test_id,graph_size=500,d=0.1)
graph.user_graph_dfs(graph.user_id) #生成图

#第一部分，只要求提供图结构
#计算图的各种统计指标，这里将其转化成了无向无权图
graph.save_graph()
# degree_hist(graph.G.to_undirected())
# print(graph.G.nodes())
# print(graph.G.to_undirected().nodes())
# graph_stat(graph.G.to_undirected())
#
#
# #第二部分，基于图的推荐，需要提供 推荐对象，步长，和推荐列表大小
# test_k=[10,20,50]
#
# for k in test_k:
#
#     #设定实验参数
#     print("success1")
#     graph.rec_for(graph.user_id,l=2,result_size=k)
#     graph.per_pagerank()
#     graph.rec_stat()
#     print("success2")
